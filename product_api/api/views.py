from itertools import product
from django.shortcuts import render, redirect
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.http import HttpResponse

from .models import Product, agent, chat, downloadexcel, reporttoday
from .serializers import AgentSerializer, ChatSerializer, DownloadExcelSerializer, ProductSerializer, ReportTodaySerializer
import requests
from django.shortcuts import render
import json
from openpyxl import load_workbook
from json2html import *
from django.contrib.auth.models import User, auth
from django.contrib import messages
import datetime
import json
fn= None
agentName=None
# Create your views here.

"""@api_view(['GET'])
def apiOverview(request):
    api_urls = {
        'List': '/product-list/',
        'Detail View': '/product-detail/<int:id>/',
        'Create': '/product-create/',
        'Update': '/product-update/<int:id>/',
        'Delete': '/product-detail/<int:id>/',
    }
    return Response(api_urls);
"""

# Save Mobile 
@api_view(['POST'])
def saveMobile(request):
    try:
        serializer = AgentSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
        return Response(serializer.data)
    except Exception as e:
        return Response("Error in Posting the data");



@api_view(['POST'])
def reportInsert(request):
    serializer = ReportTodaySerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)

@api_view(['DELETE'])
def deleteAll(request):
    reporttodays = reporttoday.objects.all()
    reporttodays.delete()
    # serializer = DownloadExcelSerializer(downloadExcels, many=True)
    return Response("Deleted")

@api_view(['GET'])
def ShowAcAll(request):
    downloadExcels = downloadexcel.objects.values('AccountNo').distinct()
    serializer = DownloadExcelSerializer(downloadExcels, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def ShowReportAll(request): 
    reporttodays = reporttoday.objects.all()
    serializer = ReportTodaySerializer(reporttodays, many=True)
    return Response(serializer.data)
    # return Response("In Progress")

# Saving Data 

@api_view(['POST'])
def saveinsertform(request):
    try:
        serializer = DownloadExcelSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
        return Response(serializer.data)
    except Exception as e:
        return Response("Error in Posting the data");

@api_view(['GET'])
def ShowsaveAll(request):
    downloadExcels = downloadexcel.objects.all()
    serializer = DownloadExcelSerializer(downloadExcels, many=True)
    return Response(serializer.data)


@api_view(['GET'])
def ViewSaved(request, pk):
    try:
        print("pk=",pk)
        # downloadExcels = downloadexcel.objects.get(AccountNo=pk)  
        downloadExcels = downloadexcel.objects.filter(AccountNo=pk)
        serializer = DownloadExcelSerializer(downloadExcels, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response("Account doesnt exist in database");

@api_view(['GET'])
def ViewUser(request, pk):
    try:
        print("pk=",pk)
        agents = agent.objects.filter(username=pk)
        serializer = AgentSerializer(agents, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response("Account doesnt exist in database");
@api_view(['GET'])
def ViewSavedUser(request, pk):
    try:
        print("pk=",pk)
        downloadExcels = downloadexcel.objects.filter(AgentName=pk)
        serializer = DownloadExcelSerializer(downloadExcels, many=True)
        return Response(serializer.data)
    except Exception as e:
        return Response("Account doesnt exist in database");
# Product realeted - upload Excel 


@api_view(['DELETE'])
def deleteDownloadExcel(request, pk):
        downloadExcels = downloadexcel.objects.filter(AgentName=pk) 
        serializer = DownloadExcelSerializer(downloadExcels, many=True)
        downloadExcels.all().delete()
        # downloadExcels = downloadexcel.objects.filter(AccountNo=pk)
        # serializer = DownloadExcelSerializer(downloadExcels, many=True)
        # for p in products:
        #     p.delete() 
        return Response("Deleted from downloadExcel")
@api_view(['DELETE'])
def deleteProduct1(request, pk):
        products = Product.objects.filter(AgentName=pk) 
        serializer = ProductSerializer(products, many=True)
        products.all().delete()
        # downloadExcels = downloadexcel.objects.filter(AccountNo=pk)
        # serializer = DownloadExcelSerializer(downloadExcels, many=True)
        # for p in products:
        #     p.delete() 
        return Response("Deleted from api_product")

@api_view(['GET'])
def ShowAll(request):
    products = Product.objects.all()
    serializer = ProductSerializer(products, many=True)
    return Response(serializer.data)  

@api_view(['GET'])
def ShowChat(request):
    chats = chat.objects.all()
    serializer = ChatSerializer(chats, many=True)
    return Response(serializer.data) 


@api_view(['POST'])
def chatLoginUpdate(request,pk):
    agents=agent.objects.get(username=pk)
    serializer = AgentSerializer(instance=agents, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)

@api_view(['GET'])
def ViewProduct(request, pk):
    try:
        product = Product.objects.get(AccountNo=pk)
        serializer = ProductSerializer(product, many=False)
        return Response(serializer.data)
    except Exception as e:
        content="Account doesnt exist in database"
        return Response(content, status=status.HTTP_404_NOT_FOUND);



@api_view(['POST'])
def CreateProduct(request):
    serializer = ProductSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)
    #     return Response(serializer.data, status=status.HTTP_201_CREATED)
    # else:
    #     return Response(serializer.data, status=status.HTTP_400_BAD_REQUEST) Createchat

@api_view(['POST'])
def Createchat(request):
    serializer = ChatSerializer(data=request.data)

    if serializer.is_valid():
        serializer.save()

    return Response(serializer.data)


@api_view(['POST'])
def updateProduct(request, pk):
    product = Product.objects.get(id=pk)
    serializer = ProductSerializer(instance=product, data=request.data)
    if serializer.is_valid():
        serializer.save()
        return Response(serializer.data)
    # else: , status=status.HTTP_201_CREATED
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['DELETE'])
def deleteProduct(request, pk):
    # reporttodays = reporttoday.objects.all()
    # reporttodays.delete()
    product = reporttoday.objects.get(AccountNo=pk)
    product.delete()

    return Response('Items delete successfully!')

@api_view(['GET'])
def deleteReport(request, pk):
    try:
        product = reporttoday.objects.get(AccountNo=pk)
        product.delete()

        return Response('Items delete successfully!')
    except Exception  as e:
        return Response("No Data");



def regform(request):
    return render(request,'registration.html')

def saveinsert(request):
    if request.method == 'POST':

        print('In Save insert form',request.POST)
        AccountNo=request.POST['AccountNo']
        AccountType=request.POST['AccountType']
        Name=request.POST['Name']
        PreviousBalance=request.POST['PreviousBalance']
        OpenDate=request.POST['OpenDate']
        Mobile=request.POST['Mobile']
        AgentName = request.POST['AgentName']
        Recieved = request.POST['Recieved']
        PresentBalance=(float(PreviousBalance)+float(Recieved))
        x = str(datetime.datetime.now())
        dt=x.split(" ")
        print(dt[0])
        time= dt[1].split(".")
        print(time[0])

        Today = dt[0]
        Time=time[0]

        data={
            "AccountNo": int(AccountNo),
            "AccountType": AccountType, 
            "Name":Name, 
            "PreviousBalance": float(PreviousBalance), 
            "OpenDate": OpenDate, 
            "Mobile": Mobile, 
            "AgentName":AgentName, 
            "Recieved": float(Recieved),
            "PresentBalance":PresentBalance,
            "Today":Today,
            "Time":Time
            
        }
        data_show=json.dumps(data, indent = 20) 
        print("data=",data_show)
        headers={'Content-Type':'application/json'}
        read=requests.post('http://117.220.197.82:80/save-insert/',json=data,headers=headers)
        print('post status=',read)
        return render(request,'saveRecieved.html')
    else:
        return render(request,'saveRecieved.html')


def uploadExcel(request):
    if "GET" == request.method:
        check=2
        return render(request, 'uploadExcel.html', {})
    else:
        print("We will delete and then upload")
        checkDelete=0
        global agentName
        print("agentName=",agentName)
        try:
            url='http://117.220.197.82/del-product/'+agentName
            x = requests.delete(url)
            print('deleted from api product db')
            url='http://117.220.197.82/del-downloadexcel/'+agentName
            print('deleted from api downloadexcel db')
            y= requests.delete(url)
            checkDelete=1
        except Exception  as e:
            checkDelete=0
        if checkDelete==1 :
            
            print('in Upload Excel agent Name',agentName)
            check=0
            try:
                messages.info(request,'Please wait data is getting inserted')
                excel_file = request.FILES["excel_file"]
                print("excel_file",excel_file)

                wb = load_workbook(excel_file)

                # getting a particular sheet by name out of many sheets
                worksheet = wb["Sheet1"]
                print(worksheet)

                excel_data = list()
                # iterating over the rows and
                # getting value from each cell in row
                for row in worksheet.iter_rows():
                    row_data = list()
                    for cell in row:
                        row_data.append(str(cell.value))
                    excel_data.append(row_data)

                # Upload to db through API
                print("No of rows=",worksheet.max_row)
                
                for i in range(1,(worksheet.max_row)):
                    print(excel_data[i][0],excel_data[i][1],excel_data[i][2],excel_data[i][3],excel_data[i][4],excel_data[i][5],)
                    data={
                        "AccountNo": excel_data[i][0],"AccountType":excel_data[i][1],"Name":excel_data[i][2],"PreviousBalance":excel_data[i][3],
                        "OpenDate":excel_data[i][4],"Mobile":excel_data[i][5],"AgentName":agentName,"id":1
    ,
                    }
                
                    headers={'Content-Type':'application/json'}
                    read=requests.post('http://117.220.197.82:80/product-create/',json=data,headers=headers)                           
                check=1
                checkdata={
                    "check":check
                }
            
            except Exception  as e:
                check=2
                checkdata={
                    "check":check,
                    "message":"Failed to insert , Please check File type"
                }
        
            global fn
            data1={
                    "is_admin":0,
                    "is_login":1,
                    "first_name":fn,
                    "check":check,
                    }
            
            return render(request,'index.html',data1)
        else:
             check=2
             data1={
                    "is_admin":0,
                    "is_login":1,
                    "first_name":fn,
                    "check":check,
                    }

             return render(request,'index.html',data1)
    


def downloadExcel1(request):
    if "GET" == request.method:
        check=2
        return render(request, 'downloadExcel.html', {})
    else:
        
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

        x = str(datetime.datetime.now())
        dt=x.split(" ")
        print(dt[0])
        time= dt[1].split(".")
        print(time[0])

        Today1 = dt[0]
        Time=time[0]

        import openpyxl
        
        wb = openpyxl.Workbook()
        sheet = wb.active
        read=requests.get('http://117.220.197.82:80/saved-list/').json()
        print("Length of Json data =",len(read))
        c1heading= sheet.cell(row=1,column=1)
        c1heading.value="Date"
        c2heading= sheet.cell(row=1,column=2)
        c2heading.value="AccountNo"
        c3heading= sheet.cell(row=1,column=3)
        c3heading.value="Amount"
        c4heading= sheet.cell(row=1,column=4)
        c4heading.value="Receipt No"
        c5heading= sheet.cell(row=1,column=5)
        c5heading.value="OpenDate"
        j=1
        for i in range(0,len(read)):
            c1 = sheet.cell(row = j+1, column = 1)
            c1.value = Today1
            
            c2 = sheet.cell(row= j+1 , column = 2)
            c2.value = int(read[i]['AccountNo'])

            c3 = sheet.cell(row= j+1 , column = 3)
            c3.value = float(read[i]['Recieved'])
            
            c4 = sheet.cell(row= j+1 , column = 4)
            c4.value = read[i]['id']
            # c4.value =  float("{:.2f}".format(float(read[i]['PresentBalance']))) 
            # c4.value = round(float(read[i]['PresentBalance']))

            # c5= sheet.cell(row= j+1 , column = 5)
            # c5.value= read[i]['OpenDate']
            j=j+1
        wb.save(response)
        
        print(read[0]['Name'])
        return response
        # return render(request, 'downloadExcel.html', {})

def readData(request):
    if "GET" == request.method:
        return render(request, 'readData.html', {})
    else:

        read=requests.get('http://117.220.197.82:80/api/product-list/').json()
        # print(read)
        
        done = json2html.convert(json=read)
        # print(done)
        with open(".//api//templates//dataview.html",'w') as f:
            f.writelines(done)

    return render(request, 'dataview.html')

def index(request):
    return render(request, "index.html")

def index1(request):
    return render(request, "index1.html")
def privatePolicy(request):
    return render(request,"privacypolicy.html")
def login1(request):
    if request.method== 'POST':
        username = request.POST['username']
        password = request.POST['password']

        user = auth.authenticate(username=username,password=password)
        print('user',user)
        global agentName
        agentName=username
        print('in Login agent Name',agentName)
        if user is not None:
            auth.login(request, user)
            return redirect("/")
        else:
            messages.info(request,'invalid credentials')
            return redirect('login')

    else:
        return render(request,'login.html')    

def login(request):
    if request.method== 'POST':
        username = request.POST['username']
        password = request.POST['password']
        
        read=requests.get('http://117.220.197.82:80/user-detail/'+(username)).json()
        print("username pwd read=",read[0]['is_admin']);
        print('type=',type(read[0]['is_admin']))
        if password==read[0]['password']:
            if read[0]['is_admin']==str(1):
                data={
                "is_admin":1,
                "is_login":1,
                "first_name":read[0]['first_name']
                }
                print(data)
                
            else:
                data={
                "is_admin":0,
                "is_login":1,
                "first_name":read[0]['first_name']
                }
            global agentName
            agentName = read[0]['username']
            global fn
            fn=read[0]['first_name']
            print("Agent Name in Login",agentName)
            # return redirect("/",data)
            return render(request,'index.html',data)
        else:
            messages.info(request,'invalid credentials')
            data={
                "is_login":0,
                "first_name":""
            }
            return render(request,'login.html',data)

    else:
        return render(request,'login.html')    

def register(request):

    if request.method == 'POST':
        first_name = request.POST['first_name']
        last_name = request.POST['last_name']
        username = request.POST['username']
        password1 = request.POST['password1']
        password2 = request.POST['password2']
        email = request.POST['email']
        mobile =  request.POST['mobile']
        if password1==password2:
            # if User.objects.filter(username=username).exists():
            #     messages.info(request,'Username already registered')
            #     return redirect('register')
            # elif User.objects.filter(email=email).exists():
            #     messages.info(request,'Email already registered')
            #     return redirect('register')
            # else:   
            #     user = User.objects.create_user(username=username, password=password1, email=email,first_name=first_name,last_name=last_name)
            #     user.save();
            data={  "first_name":first_name,
                    "last_name": last_name,
                    "username":username,
                     "email":email,
                    "mobile":mobile,
                    "password":password1,
                }
            print("data=",data)
            headers={'Content-Type':'application/json'}
            read=requests.post('http://117.220.197.82:80/save-mobile/',json=data,headers=headers)
            print(read)
            print('user created')
            return redirect('login')
            # global fn
            # data1={
            #         "first_name":fn,
            #         "is_login":1
            # }
            # return render(request,'login.html')

        else:
            messages.info(request,'password not matching..')    
            return redirect('register')
        return redirect('/')
        
    else:
        return render(request,'register.html')



def logout(request):
    auth.logout(request)
    return redirect('/')    

def saveform(request):
    return render(request,'saveRecieved.html')


def saverecieved1(request):
     if request.method == 'POST':
        x = str(datetime.datetime.now())
        dt=x.split(" ")
        print(dt[0])
        time= dt[1].split(".")
        print(time[0])

        Today = dt[0]
        Time=time[0]
        data={}
        AccountNo = request.POST['AccountNo']
        print("================ After searching Account =============\n")
        readSave=requests.get('http://117.220.197.82:80/saved-detail/'+(AccountNo)).json()
        print ("got details=",readSave)
        if readSave=="Account doesnt exist in database" or readSave==[]:
            print("================ Getting from upload Excel  =============\n")
            read=requests.get('http://117.220.197.82:80/product-detail/'+(AccountNo)).json()
            print('read=',read)
            data={
                "AccountNo": int(read['AccountNo']),"AccountType":read['AccountType'],"Name":read['Name'],
                "PreviousBalance":float(read['PreviousBalance']),
                        "OpenDate":read['OpenDate'],"Mobile":read['Mobile'],"AgentName":read['AgentName'],
            }
        else:
            # print('readSave[AccountNo]=',readSave[-1]['PresentBalance'])
            print("================ Getting from Saved Table  =============\n")
            read=readSave
            
            data={
                "AccountNo": int(read[-1]['AccountNo']),"AccountType":read[-1]['AccountType'],"Name":read[-1]['Name'],
                "PreviousBalance":float(read[-1]['PresentBalance']),
                        "OpenDate":read['OpenDate'],"Mobile":read[-1]['Mobile'],"AgentName":read[-1]['AgentName'],
            }
           
        print('data=',data)    
        # headers={'Content-Type':'application/json'}
        # getPostResult=requests.post('http://117.220.197.82:80/save-insert/',json=data,headers=headers)
        # print("getPostResult=",getPostResult)
     return render(request,'saveRecieved.html',data)


def saveReport12(request):
     if request.method == 'POST':
        readSave=requests.get('http://117.220.197.82:80/unique-ac-list/').json()
        if(len(readSave)>0):
            for i in range(0,len(readSave)):
                    AccountNo = readSave[i]['AccountNo']
                    deleteReport = requests.get('http://117.220.197.82:80/report-delete/'+(AccountNo)).json()
                    print("del status=",deleteReport)
        

        data={}
        print('length of readSave=',len(readSave))
        if( len(readSave)>0):
            total=[]
            for i in range(0,len(readSave)):
                AccountNo = readSave[i]['AccountNo']
                read = requests.get('http://117.220.197.82:80/saved-detail/'+(AccountNo)).json()
                print('length of read=',len(read))
                sum=0
                
                for i in range(0,len(read)):
                    sum=float(sum)+float(read[i]['Recieved'])
                total.append(sum)
                print('Total Recieved=',sum)
            print('Total length of  List =',len(total))
        # print('REport AccountNo=',len(readSave))
        if( len(readSave)>0):
            for i in range(0,len(readSave)):
                AccountNo = readSave[i]['AccountNo']
                read = requests.get('http://117.220.197.82:80/saved-detail/'+(AccountNo)).json()
                
                # print('read=',read[-1])
                data={
                    "AccountNo": read[-1]['AccountNo'],"AccountType":read[-1]['AccountType'],"Name":read[-1]['Name'],
                    "Recieved":total[i],
                            "OpenDate":read[-1]['OpenDate'],"Mobile":read[-1]['Mobile'],"AgentName":read[-1]['AgentName'],
                }
                # print('data=',data)
                headers={'Content-Type':'application/json'}
                getPostResult=requests.post('http://117.220.197.82:80/report-insert/',json=data,headers=headers)
                # print("getPostResult=",getPostResult)
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

            import openpyxl
            
            wb = openpyxl.Workbook()
            sheet = wb.active
            read=requests.get('http://117.220.197.82:80/report-list/').json()
            print("Length of Json data =",len(read))
            c1heading= sheet.cell(row=1,column=1)
            c1heading.value="Date"
            c2heading= sheet.cell(row=1,column=2)
            c2heading.value="AccountNo"
            c3heading= sheet.cell(row=1,column=3)
            c3heading.value="Amount"
            c4heading= sheet.cell(row=1,column=4)
            c4heading.value="RecieptNo"
           
            j=1
            for i in range(0,len(read)):
                c1 = sheet.cell(row = j+1, column = 1)
                c1.value = "2022-08-25"
                
                c2 = sheet.cell(row= j+1 , column = 2)
                c2.value = int(read[i]['AccountNo'])

                c3 = sheet.cell(row= j+1 , column = 3)
                c3.value = float(read[i]['Recieved'])
                
                c4 = sheet.cell(row= j+1 , column = 4)
                c4.value = read[i]['id']
               
                j=j+1
            wb.save(response)
            
            return response
        else:
            return Response("No Data")


def saveReport1(request):
     if request.method == 'POST':
        global agentName
        print("inside the download report agentName= ",agentName)       
        x = str(datetime.datetime.now())
        dt=x.split(" ")
        print(dt[0])
        time= dt[1].split(".")
        print(time[0])
        d= datetime.datetime.today()
        Today = d.strftime("%d-%m-%Y") 

        Time=time[0]
        read = requests.get('http://117.220.197.82:80/saved-detail-user/'+(agentName)).json()
        print('length of read=',len(read))
        data={
                    "AccountNo": read[-1]['AccountNo'],"AccountType":read[-1]['AccountType'],"Name":read[-1]['Name'],
                    "Recieved":read[-1]['Recieved'],
                            "OpenDate":read[-1]['OpenDate'],"Mobile":read[-1]['Mobile'],"AgentName":read[-1]['AgentName'],
                }
               
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="report.xlsx"'

            
        import openpyxl
        wb = openpyxl.Workbook()
        sheet = wb.active
        # read=requests.get('http://117.220.197.82:80/report-list/').json()
        print("Length of Json data =",len(read))
        c1heading= sheet.cell(row=1,column=1)
        c1heading.value="Date"
        c2heading= sheet.cell(row=1,column=2)
        c2heading.value="AccountNo"
        c3heading= sheet.cell(row=1,column=3)
        c3heading.value="Amount"
        c4heading= sheet.cell(row=1,column=4)
        c4heading.value="RecieptNo"
           
        j=1
        for i in range(0,len(read)):
                c1 = sheet.cell(row = j+1, column = 1)
                c1.value = Today
                
                c2 = sheet.cell(row= j+1 , column = 2)
                c2.value = int(read[i]['AccountNo'])

                c3 = sheet.cell(row= j+1 , column = 3)
                c3.value = float(read[i]['Recieved'])
                
                c4 = sheet.cell(row= j+1 , column = 4)
                c4.value = read[i]['id']
               
                j=j+1
        wb.save(response)
            
        return response
       
